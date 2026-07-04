#!/usr/bin/env python3
"""Read-only collector for SB Reports > Messenger Pages zero-delivery suspects.

Plan source: Rodolfo 2026-07-03.
- full scope: digital-trust + digital-trust-2 active publishers
- report route: https://app.smartbiddingdigital.com/reports/messenger
- operational filter: page sent broadcast but delivered zero messages
  (bd_sends>0 and bd_delivereds==0 / bd_delivered_rate==0)
- also exports active Account/Page restrictions for DTR reconference

No writes to SmartBidding.
"""
import argparse, asyncio, csv, datetime as dt, importlib.util, json, pathlib, re
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

BASE = pathlib.Path('/root/mgs-agent')
NY = ZoneInfo('America/New_York')

spec = importlib.util.spec_from_file_location('dtrsync', BASE / 'scripts/dtr-sb-page-health-sync.py')
assert spec and spec.loader
sync = importlib.util.module_from_spec(spec)
spec.loader.exec_module(sync)


def now_et():
    return dt.datetime.now(NY)


def iso_z(d: dt.datetime) -> str:
    return d.astimezone(dt.timezone.utc).isoformat(timespec='seconds').replace('+00:00', 'Z')


def norm(v):
    return '' if v is None else str(v).strip()


def get_ci(row, *names):
    if not isinstance(row, dict):
        return ''
    lower = {str(k).lower(): k for k in row.keys()}
    for name in names:
        k = lower.get(str(name).lower())
        if k is not None:
            return row.get(k)
    return ''


def num(v):
    if v is None or v == '':
        return 0.0
    s = str(v).strip().replace('%','').replace(',','.')
    s = re.sub(r'[^0-9.\-]+', '', s)
    if not s or s == '-':
        return 0.0
    try:
        return float(s)
    except Exception:
        return 0.0


def delivered_rate(row):
    raw = get_ci(row, 'BD_DELIVERED_RATE', 'bd_delivered_rate', 'BD DELIVERED RATE')
    if raw not in (None, ''):
        val = num(raw)
        return val / 100 if val > 1 else val
    sends = num(get_ci(row, 'BD_SENDS', 'bd_sends'))
    delivered = num(get_ci(row, 'BD_DELIVEREDS', 'bd_delivereds', 'BD_DELIVERED', 'DELIVEREDS'))
    return delivered / sends if sends else 0.0


def safe_row(row):
    bd_sends = num(get_ci(row, 'BD_SENDS', 'bd_sends'))
    bd_delivereds = num(get_ci(row, 'BD_DELIVEREDS', 'bd_delivereds', 'DELIVEREDS'))
    rate = delivered_rate(row)
    return {
        'company': get_ci(row, 'COMPANY', 'company'),
        'domain': get_ci(row, 'DOMAIN', 'domain'),
        'user_login': get_ci(row, 'USER_LOGIN', 'LOGIN', 'login', 'user'),
        'profile_name': get_ci(row, 'PROFILE_NAME', 'profile_name', 'segurador'),
        'page_id': get_ci(row, 'PAGE_ID', 'page_id'),
        'fb_page_id': get_ci(row, 'FB_PAGE_ID', 'fb_page_id'),
        'page_name': get_ci(row, 'PAGE_NAME', 'page_name'),
        'status': get_ci(row, 'STATUS', 'status'),
        'bd_sends': int(bd_sends) if bd_sends.is_integer() else bd_sends,
        'bd_delivereds': int(bd_delivereds) if bd_delivereds.is_integer() else bd_delivereds,
        'bd_delivered_rate': round(rate, 4),
        'sends': get_ci(row, 'SENDS', 'sends'),
        'delivereds': get_ci(row, 'DELIVEREDS', 'delivereds'),
        'leads': get_ci(row, 'LEADS', 'leads'),
        'revenue': get_ci(row, 'REVENUE', 'revenue'),
        'raw_keys': ','.join(sorted(str(k) for k in row.keys())[:80]),
    }


def write_csv(path, rows):
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(rows[0].keys()) if rows else ['empty']
    with path.open('w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_xlsx(path, sheets):
    if not Workbook:
        return None
    wb = Workbook()
    first = True
    for title, rows in sheets.items():
        ws = wb.active if first else wb.create_sheet()
        first = False
        ws.title = title[:31]
        fields = list(rows[0].keys()) if rows else ['empty']
        ws.append(fields)
        for c in ws[1]:
            c.font = Font(bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor='1F4E78')
        for r in rows:
            ws.append([r.get(f, '') for f in fields])
        for i, f in enumerate(fields, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(len(f)+3, 14), 42)
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    return str(path)


async def fetch_report_rows(ctx, h, pubs, initial, final):
    payload = {'initialDate': iso_z(initial), 'finalDate': iso_z(final), 'publishers': pubs, 'currency': None}
    r = await ctx.request.post('https://api.jbfdigital.com.br/report/messenger', headers=h, data=json.dumps(payload), timeout=180000)
    txt = await r.text()
    try:
        data = json.loads(txt) if txt else None
    except Exception:
        raise RuntimeError(f'report/messenger returned non-json status={r.status} body={txt[:300]}')
    if r.status not in (200, 201):
        raise RuntimeError(f'report/messenger status={r.status} body={txt[:300]}')
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        for k in ('data', 'rows', 'result', 'items'):
            if isinstance(data.get(k), list):
                return data[k]
    raise RuntimeError(f'unexpected report/messenger payload type={type(data).__name__}')


async def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--days', type=int, default=1, help='ET days ending now; default today/last 1 day window')
    ap.add_argument('--initial-et', default='', help='YYYY-MM-DD start in America/New_York')
    ap.add_argument('--final-et', default='', help='YYYY-MM-DD end in America/New_York inclusive')
    ap.add_argument('--out-prefix', default='')
    ap.add_argument('--legacy-low-delivery', action='store_true', help='Legacy/debug only: bd_sends>0 and bd_delivered_rate<0.5. Do not use for Rodolfo page-health plan.')
    args = ap.parse_args()

    end = now_et()
    if args.final_et:
        d = dt.date.fromisoformat(args.final_et)
        end = dt.datetime.combine(d, dt.time.max, tzinfo=NY).replace(microsecond=0)
    if args.initial_et:
        d = dt.date.fromisoformat(args.initial_et)
        start = dt.datetime.combine(d, dt.time.min, tzinfo=NY)
    else:
        start = (end - dt.timedelta(days=max(args.days, 1))).replace(hour=0, minute=0, second=0, microsecond=0)

    stamp = now_et().strftime('%Y%m%d-%H%M%S')
    prefix = args.out_prefix or str(BASE / 'reports' / f'sb-messenger-low-delivery-{stamp}')
    p = browser = None
    summary = {'ok': True, 'started_at_et': now_et().isoformat(timespec='seconds'), 'initial_et': start.isoformat(), 'final_et': end.isoformat()}
    try:
        p, browser, ctx, h = await sync.get_sb_context()
        pubs, campaign_rows = await sync.fetch_sb_rows(ctx, h)
        report_rows = await fetch_report_rows(ctx, h, pubs, start, end)
        zero = []
        for row in report_rows:
            sends = num(get_ci(row, 'BD_SENDS', 'bd_sends'))
            delivereds = num(get_ci(row, 'BD_DELIVEREDS', 'bd_delivereds', 'BD_DELIVERED', 'DELIVEREDS'))
            rate = delivered_rate(row)
            if args.legacy_low_delivery:
                match = sends > 0 and rate < 0.5
            else:
                # Rodolfo-corrected operational rule: page sent broadcast and delivered ZERO messages.
                # Use delivered count as primary; accept rate==0 fallback when delivered count is absent.
                delivered_raw = get_ci(row, 'BD_DELIVEREDS', 'bd_delivereds', 'BD_DELIVERED', 'DELIVEREDS')
                match = sends > 0 and (delivereds == 0 or (delivered_raw in (None, '') and rate == 0))
            if match:
                zero.append(safe_row(row))
        tday = now_et().date().isoformat()
        restricted = [sync.public_row(r) for r in campaign_rows if sync.active_restricted(r, tday)]
        filter_label = 'bd_sends>0 bd_delivered=0' if not args.legacy_low_delivery else 'LEGACY bd_sends>0 bd_delivered_rate<0.5'
        summary.update({
            'publishers': len(pubs),
            'campaign_rows': len(campaign_rows),
            'report_rows': len(report_rows),
            'zero_delivery_rows': len(zero),
            'active_restricted_rows': len(restricted),
            'filter': filter_label,
        })
        out_json = pathlib.Path(prefix + '.json')
        out_zero_csv = pathlib.Path(prefix + '-zero-delivery.csv')
        out_rest_csv = pathlib.Path(prefix + '-active-restricted.csv')
        out_xlsx = pathlib.Path(prefix + '.xlsx')
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(json.dumps({'summary': summary, 'zero_delivery': zero, 'active_restricted': restricted}, ensure_ascii=False, indent=2), encoding='utf-8')
        write_csv(out_zero_csv, zero)
        write_csv(out_rest_csv, restricted)
        write_xlsx(out_xlsx, {'Zero delivery': zero, 'Active restricted': restricted, 'Resumo': [{'campo': k, 'valor': v} for k, v in summary.items()]})
        summary.update({'json': str(out_json), 'zero_delivery_csv': str(out_zero_csv), 'active_restricted_csv': str(out_rest_csv), 'xlsx': str(out_xlsx)})
    except Exception as exc:
        summary['ok'] = False
        summary['error'] = f'{type(exc).__name__}: {exc}'
    finally:
        try:
            if browser: await browser.close()
            if p: await p.stop()
        except Exception:
            pass
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    raise SystemExit(0 if summary.get('ok') else 2)


if __name__ == '__main__':
    asyncio.run(main())
