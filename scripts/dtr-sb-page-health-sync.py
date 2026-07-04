#!/usr/bin/env python3
"""DTR -> SmartBidding page health sync.

Validated workflow:
- active bot users from migration Sheet only;
- DigitalTRChat account/segurador -> real page selector -> search_page_id per page;
- latest Completed report only;
- update SB NOTES for every non-Sent result, all SB statuses;
- apply/clear RESTRICTED_UNTIL only with validated rules;
- Blocked -> Broadcast only if Facebook page opens normally.
"""
import argparse, asyncio, csv, html, io, json, os, re, subprocess, sys, tempfile, urllib.parse, urllib.request
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo
from playwright.async_api import async_playwright

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except Exception:
    Workbook = None

BASE_DIR=Path('/root/mgs-agent')
SHEET_ID='1sTkBE6RQPQ3obq1j6m8RSu_22beEUbZjkQ-OttI01XY'
MIGRATION_GID='562940072'
DTR_BASE='https://digitaltrchat.com'
SB_STATE='/tmp/smartbidding_state_headed.json'
NY=ZoneInfo('America/New_York')
LOG_DIR=BASE_DIR/'logs'
REPORT_DIR=BASE_DIR/'reports'
STATE_PATH=BASE_DIR/'data/dtr-sb-page-health-sync-state.json'

MONTHS_EN={'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,'july':7,'august':8,'september':9,'october':10,'november':11,'december':12}


def norm(v): return '' if v is None else str(v).strip()
def norm_email(v): return norm(v).lower()
def clean(v): return html.unescape(re.sub(r'<[^>]+>',' ',str(v or ''))).replace('\u202f',' ').replace('\xa0',' ').strip()
def today(): return datetime.now(NY).date().isoformat()
def now_iso(): return datetime.now(NY).isoformat(timespec='seconds')
def date_only(v): return norm(v)[:10]

def op(cmd, timeout=30): return subprocess.check_output(cmd, text=True, env=os.environ.copy(), timeout=timeout).strip()
def op_json(cmd): return json.loads(op(cmd, timeout=60))

def sheet_rows():
    url=f'https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid={MIGRATION_GID}'
    data=urllib.request.urlopen(url, timeout=60).read().decode('utf-8-sig')
    return list(csv.DictReader(io.StringIO(data)))

def active_users_from_sheet(rows):
    users=[]
    for r in rows:
        u=norm_email(r.get('User'))
        if '@' not in u: continue
        if not norm(r.get('NO APP')): continue
        if norm(r.get('Removidos acumulado')).upper()=='X': continue
        users.append(u)
    return sorted(set(users))

def discover_dtr_items(target_users):
    vault=os.environ.get('OP_DEFAULT_VAULT','MGS Conteúdo')
    items=op_json(['op','item','list','--vault',vault,'--format','json'])
    titles=[i.get('title','') for i in items if norm(i.get('title')).lower().startswith('digitaltrchat - disparos')]
    matched={}; errors=[]
    for t in sorted(set(titles), key=str.lower):
        try: u=op(['op','item','get',t,'--vault',vault,'--fields','username','--reveal']).lower()
        except Exception as exc:
            errors.append({'item':t,'error':type(exc).__name__}); continue
        if u in target_users and u not in matched: matched[u]=t
    return matched, sorted(set(target_users)-set(matched)), errors

def op_password(item):
    vault=os.environ.get('OP_DEFAULT_VAULT','MGS Conteúdo')
    for f in ('credential','password'):
        try: return op(['op','item','get',item,'--vault',vault,'--fields',f,'--reveal'])
        except Exception: pass
    raise RuntimeError(f'password field not found for {item}')

def parse_restricted_date(text, year=None):
    t=clean(text); y=year or datetime.now(NY).year
    m=re.search(r'until\s+([A-Za-z]+)\s+(\d{1,2})\s+at\s+(\d{1,2}):(\d{2})\s*([AP]M)', t, re.I)
    if m:
        mon=MONTHS_EN.get(m.group(1).lower()); day=int(m.group(2)); hh=int(m.group(3)); mm=int(m.group(4)); ap=m.group(5).upper()
        if mon:
            if ap=='PM' and hh!=12: hh+=12
            if ap=='AM' and hh==12: hh=0
            return f'{y:04d}-{mon:02d}-{day:02d}', f'{y:04d}-{mon:02d}-{day:02d} {hh:02d}:{mm:02d}'
    return None, None

def date_from_text(text):
    t=clean(text)
    m=re.search(r'\b(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d{1,2},\s*\d{2}\s+\d{1,2}:\d{2}\b', t, re.I)
    return m.group(0) if m else ''

def classify_report(raw):
    t=clean(raw); low=t.lower(); codes=[]
    checks=[('#2022',['#2022','temporarily restricted','restring']),('#10',['#10','outside of allowed window','fora do espaço de tempo permitido','fuera del período permitido']),('#551',['#551',"isn't available",'não está disponível','no se encuentra disponible']),('#100',['#100','missing one or more params','no matching user found','não foi possível encontrar o modelo','no se puede encontrar la plantilla']),('APP_DELETED',['application has been deleted','aplicativo foi excluído']),('PERMISSION',['pages_messaging permission','permission(s) must be granted','before impersonatin']),('TOKEN',['oauth','token','session'])]
    for code, needles in checks:
        if any(n.lower() in low for n in needles): codes.append(code)
    ru,rut=parse_restricted_date(t)
    if not codes:
        if re.search(r'\bSent\b|Enviado|Delivered|Entregado', t, re.I):
            return {'status':'SENT','codes':[], 'note_code':'', 'restricted_until':None, 'restricted_until_time':None, 'raw':t[:1200]}
        return {'status':'SEM_COMPLETED' if not t else 'OTHER', 'codes':([] if not t else ['OTHER']), 'note_code':('SEM_COMPLETED' if not t else 'OTHER'), 'restricted_until':None, 'restricted_until_time':None, 'raw':t[:1200]}
    return {'status':'ERROR','codes':codes,'note_code':' - '.join(codes),'restricted_until':ru,'restricted_until_time':rut,'raw':t[:1200]}

def campaign_form(csrf, page_id, length=10):
    form={'draw':'1','start':'0','length':str(length),'search_page_id':str(page_id),'search_value':'','search_status':'2','campaign_date_range':'','csrf_token':csrf,'order[0][column]':'12','order[0][dir]':'desc','search[value]':'','search[regex]':'false'}
    for i in range(14):
        form[f'columns[{i}][data]']=str(i); form[f'columns[{i}][searchable]']='true'; form[f'columns[{i}][orderable]']='true'; form[f'columns[{i}][search][value]']=''; form[f'columns[{i}][search][regex]']='false'
    return form

def report_form(csrf, campaign_id, length=100):
    form={'draw':'1','start':'0','length':str(length),'campaign_id':str(campaign_id),'csrf_token':csrf,'order[0][column]':'3','order[0][dir]':'desc','search[value]':'','search[regex]':'false'}
    for i in range(9):
        form[f'columns[{i}][data]']=str(i); form[f'columns[{i}][searchable]']='true'; form[f'columns[{i}][orderable]']='true'; form[f'columns[{i}][search][value]']=''; form[f'columns[{i}][search][regex]']='false'
    return form

async def dtr_post_json(ctx, url, form, ref):
    r=await ctx.request.post(url, form=form, headers={'X-Requested-With':'XMLHttpRequest','Referer':ref}, timeout=60000)
    txt=await r.text()
    try: return json.loads(txt) if txt else {}
    except Exception: return {'_parse_error':txt[:500], '_status':r.status}

def campaign_id_from_row(row):
    for cell in row:
        m=re.search(r"cam-id=['\"]?(\d+)", str(cell))
        if m: return m.group(1)
    return ''

def fb_id_from_row(row):
    txt=' '.join(str(x) for x in row)
    for pat in [r'facebook\.com\\?/(\d+)', r'facebook\.com/(\d+)', r'facebook\.com%2F(\d+)']:
        m=re.search(pat, txt)
        if m: return m.group(1)
    return ''

async def scan_dtr_user(username, item, limit_accounts=0, limit_pages=0):
    out={'username':username,'accounts':[], 'reports':[], 'errors':[], 'login_ok':False}
    try:
        password=op_password(item)
    except Exception as exc:
        out['errors'].append(f'credential_error: {type(exc).__name__}: {exc}')
        return out
    async with async_playwright() as p:
        browser=await p.chromium.launch(headless=True,args=['--no-sandbox'])
        ctx=await browser.new_context(viewport={'width':1600,'height':1000})
        page=await ctx.new_page(); url=f'{DTR_BASE}/messenger_bot_enhancers/subscriber_broadcast_campaign'
        try:
            await page.goto(f'{DTR_BASE}/home/login', wait_until='domcontentloaded', timeout=60000)
            inputs=page.locator('input:visible'); await inputs.nth(0).fill(username); await inputs.nth(1).fill(password)
            await page.locator('button:visible, input[type=submit]:visible').last.click(); await page.wait_for_timeout(3500)
            await page.goto(url, wait_until='domcontentloaded', timeout=60000)
            csrf=await page.locator('#csrf_token').input_value(timeout=10000); out['login_ok']=True
            accs=await page.evaluate("""() => Array.from(document.querySelectorAll('.account_switch')).map(el=>({id:el.getAttribute('data-id')||el.dataset.id||'', name:(el.innerText||el.textContent||'').trim()})).filter(x=>x.id||x.name)""")
            if not accs: accs=[{'id':'','name':'default'}]
            seen=set(); uniq=[]
            for a in accs:
                k=(a.get('id','')+'|'+a.get('name','')).strip()
                if k not in seen: seen.add(k); uniq.append(a)
            if limit_accounts: uniq=uniq[:limit_accounts]
            signatures=[]
            for a in uniq:
                aid=a.get('id') or ''; aname=clean(a.get('name') or 'default') or 'default'
                acc={'id':aid,'name':aname,'pages':0,'latest_completed':0,'no_completed':0,'signature':[],'errors':[]}
                out['accounts'].append(acc)
                try:
                    if aid:
                        await ctx.request.post(f'{DTR_BASE}/social_accounts/fb_rx_account_switch', form={'id':aid,'csrf_token':csrf}, headers={'X-Requested-With':'XMLHttpRequest','Referer':url}, timeout=60000)
                        await page.goto(url, wait_until='domcontentloaded', timeout=60000); await page.wait_for_timeout(700)
                        csrf=await page.locator('#csrf_token').input_value(timeout=10000)
                    opts=await page.evaluate("""() => Array.from(document.querySelectorAll('select#search_page_id option, select[name=search_page_id] option')).map(o=>({value:o.value||'', text:(o.innerText||o.textContent||'').trim()})).filter(x=>x.value && x.value!='0' && !/select|page/i.test(x.text))""")
                    if limit_pages: opts=opts[:limit_pages]
                    acc['pages']=len(opts)
                    sig=[]
                    for opt in opts:
                        page_id=str(opt['value']); page_name=clean(opt['text'])
                        camp=await dtr_post_json(ctx, url+'_data', campaign_form(csrf,page_id), url)
                        rows=camp.get('data') or []
                        cid=''; crow=None
                        for r in rows:
                            cid=campaign_id_from_row(r)
                            if cid: crow=r; break
                        if not cid:
                            acc['no_completed']+=1
                            cls={'status':'SEM_COMPLETED','codes':[],'note_code':'SEM_COMPLETED','restricted_until':None,'restricted_until_time':None,'raw':''}
                            out['reports'].append({'bot_user':username,'account_id':aid,'account_name':aname,'dtr_page_id':page_id,'page_name':page_name,'fb_page_id':'','campaign_id':'','completed_date':'','classification':cls})
                            continue
                        acc['latest_completed']+=1; sig.append(cid)
                        rep=await dtr_post_json(ctx, f'{DTR_BASE}/messenger_bot_enhancers/campaign_sent_status_data', report_form(csrf,cid), url)
                        raw=' '.join(' '.join(str(x) for x in rr) for rr in (rep.get('data') or []))
                        cls=classify_report(raw)
                        out['reports'].append({'bot_user':username,'account_id':aid,'account_name':aname,'dtr_page_id':page_id,'page_name':page_name,'fb_page_id':fb_id_from_row(crow or []),'campaign_id':cid,'completed_date':date_from_text(raw) or date_from_text(' '.join(str(x) for x in (crow or []))),'classification':cls})
                    acc['signature']=sig[:10]; signatures.append(tuple(sig[:5]))
                except Exception as exc:
                    acc['errors'].append(f'{type(exc).__name__}: {exc}')
            out['context_signatures_unique']=len(set(signatures))
        except Exception as exc:
            out['errors'].append(f'{type(exc).__name__}: {exc}')
        finally:
            await browser.close()
    return out

async def get_sb_context():
    p=await async_playwright().start()
    browser=await p.chromium.launch(headless=False,args=['--disable-blink-features=AutomationControlled'])
    ctx=await browser.new_context(storage_state=SB_STATE, viewport={'width':1600,'height':1000}, user_agent='Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/148.0.0.0 Safari/537.36')
    page=await ctx.new_page(); headers={}
    async def on_req(req):
        if 'api.jbfdigital.com.br' in req.url:
            headers.update(await req.all_headers())
    page.on('request', on_req)
    await page.goto('https://app.smartbiddingdigital.com/accounts', wait_until='domcontentloaded', timeout=60000); await page.wait_for_timeout(5000)
    h={k:v for k,v in headers.items() if k.lower() in {'authorization','accept','content-type'}}
    h.update({'origin':'https://app.smartbiddingdigital.com','referer':'https://app.smartbiddingdigital.com/'})
    return p,browser,ctx,h

async def fetch_sb_rows(ctx,h):
    rc=await ctx.request.get('https://api.jbfdigital.com.br/company', headers=h, timeout=120000)
    companies=await rc.json(); pubs=[]
    for c in companies:
        for pub in c.get('publishers') or []:
            if pub.get('active') and pub.get('publisherId'): pubs.append(pub['publisherId'])
    qs='&'.join('companies[]='+urllib.parse.quote(x) for x in pubs)+'&source=Messenger'
    r=await ctx.request.get('https://api.jbfdigital.com.br/campaigns/Messenger?'+qs, headers=h, timeout=120000)
    rows=await r.json()
    if r.status!=200 or not isinstance(rows,list): raise RuntimeError(f'bad SB campaigns response status={r.status}')
    return pubs, rows

def build_sb_indexes(rows):
    by_fb=defaultdict(list); by_user_page=defaultdict(list); by_user_name=defaultdict(list)
    for r in rows:
        if norm(r.get('FB_PAGE_ID')): by_fb[norm(r.get('FB_PAGE_ID'))].append(r)
        if norm(r.get('USER_LOGIN')) and norm(r.get('PAGE_ID')): by_user_page[(norm_email(r.get('USER_LOGIN')), norm(r.get('PAGE_ID')))].append(r)
        if norm(r.get('USER_LOGIN')) and norm(r.get('PAGE_NAME')): by_user_name[(norm_email(r.get('USER_LOGIN')), norm(r.get('PAGE_NAME')).lower())].append(r)
    return by_fb, by_user_page, by_user_name

def match_sb(rep, indexes):
    by_fb, by_user_page, by_user_name = indexes
    cands=[]
    fb=norm(rep.get('fb_page_id'))
    if fb: cands=by_fb.get(fb, [])
    if not cands: cands=by_user_page.get((norm_email(rep.get('bot_user')), norm(rep.get('dtr_page_id'))), [])
    if not cands: cands=by_user_name.get((norm_email(rep.get('bot_user')), norm(rep.get('page_name')).lower()), [])
    if len(cands)==1: return cands[0], None
    if not cands: return None, 'no_match'
    return None, f'ambiguous_{len(cands)}'

def append_note(existing, note_code):
    existing=norm(existing)
    code=norm(note_code)
    if not code: return existing, False
    parts=[p.strip() for p in re.split(r'\s+-\s+', code) if p.strip()]
    missing=[p for p in parts if not re.search(r'(?<![\w#])'+re.escape(p)+r'(?![\w#])', existing)]
    if not missing: return existing, False
    suffix=' - '.join(missing)
    return (existing + ' - ' + suffix) if existing else suffix, True

def active_restricted(row, tday):
    ru=date_only(row.get('RESTRICTED_UNTIL'))
    return bool(ru and ru>=tday)

def public_row(r):
    return {k:r.get(k) for k in ['ID','PAGE_ID','FB_PAGE_ID','PAGE_NAME','USER_LOGIN','PROFILE_NAME','STATUS','RESTRICTED_UNTIL','NOTES','BROADCAST_TEMPLATE_NAME']}

async def fb_page_opens(ctx, fb_page_id):
    if not fb_page_id: return 'ambiguous'
    url=f'https://www.facebook.com/{fb_page_id}'
    try:
        r=await ctx.request.get(url, timeout=20000, headers={'user-agent':'Mozilla/5.0'})
        txt=(await r.text())[:200000]
        low=txt.lower()
        if "this content isn't available right now" in low or "content isn't available" in low or 'go to feed' in low:
            return 'unavailable'
        if r.status in (200,302) and ('facebook' in low) and ('page' in low or 'profile' in low or 'home_icon' in low):
            return 'available'
        return 'ambiguous'
    except Exception:
        return 'ambiguous'

async def sb_get_row(ctx,h,row_id):
    r=await ctx.request.get(f'https://api.jbfdigital.com.br/campaigns/Messenger/{row_id}', headers=h, timeout=120000)
    if r.status != 200:
        return None
    return await r.json()

async def sb_update(ctx,h,row,payload):
    """Update one SB Messenger row.

    update-many persists STATUS/RESTRICTED_UNTIL but silently ignores NOTES.
    The modal's single-row save path persists NOTES via POST /campaigns/Messenger
    when sent with the row's editable fields. Use that route whenever NOTES is
    present; otherwise use update-many for lightweight status/restriction updates.
    """
    row_id=str(row.get('ID'))
    if 'NOTES' in payload:
        current=await sb_get_row(ctx,h,row_id)
        if not current:
            return 404, 'row not found before save'
        allowed=['ID','PUBLISHER_ID','MESSENGER_USER_ID','PAGE_ID','FB_PAGE_ID','PAGE_NAME','UTM_CAMPAIGN','LEADS','STATUS','SOURCE','VERTICAL','COUNTRY','NOTES','HOLDER1','HOLDER2','ADVERTISER','DATE_START','BROADCAST_TEMPLATE_ID','BROADCAST_TIME','BROADCAST_CURRENT_MESSAGE_ID','BROADCAST_MESSAGE_ID','BROADCAST_LAST_SCHEDULE','RESTRICTED_UNTIL']
        save_payload={k:current.get(k) for k in allowed if k in current}
        notes_value=payload['NOTES']
        rest_payload={k:v for k,v in payload.items() if k!='NOTES'}
        # Save NOTES with the current row status first. Some Blocked rows 500 if
        # NOTES and STATUS=Broadcast are posted together. Also, update-many
        # ignores RESTRICTED_UNTIL=null; the single-row POST route is the proven
        # way to clear a restriction while preserving NOTES.
        save_payload['NOTES']=notes_value
        if rest_payload.get('RESTRICTED_UNTIL', 'not-present') is None:
            save_payload['RESTRICTED_UNTIL']=None
            rest_payload.pop('RESTRICTED_UNTIL', None)
        r=await ctx.request.post('https://api.jbfdigital.com.br/campaigns/Messenger', headers=h, data=json.dumps(save_payload), timeout=120000)
        txt=(await r.text())[:500]
        if r.status not in (200,201):
            return r.status, txt
        if rest_payload:
            upd={**rest_payload,'ids':[row_id]}
            r2=await ctx.request.put('https://api.jbfdigital.com.br/campaigns/Messenger/update-many', headers=h, data=json.dumps(upd), timeout=120000)
            txt2=(await r2.text())[:500]
            if r2.status not in (200,201):
                return r2.status, txt2
            return r2.status, txt2
        return r.status, txt
    if payload.get('RESTRICTED_UNTIL', 'not-present') is None:
        current=await sb_get_row(ctx,h,row_id)
        if not current:
            return 404, 'row not found before clear restriction'
        allowed=['ID','PUBLISHER_ID','MESSENGER_USER_ID','PAGE_ID','FB_PAGE_ID','PAGE_NAME','UTM_CAMPAIGN','LEADS','STATUS','SOURCE','VERTICAL','COUNTRY','NOTES','HOLDER1','HOLDER2','ADVERTISER','DATE_START','BROADCAST_TEMPLATE_ID','BROADCAST_TIME','BROADCAST_CURRENT_MESSAGE_ID','BROADCAST_MESSAGE_ID','BROADCAST_LAST_SCHEDULE','RESTRICTED_UNTIL']
        save_payload={k:current.get(k) for k in allowed if k in current}
        save_payload['RESTRICTED_UNTIL']=None
        r=await ctx.request.post('https://api.jbfdigital.com.br/campaigns/Messenger', headers=h, data=json.dumps(save_payload), timeout=120000)
        return r.status, (await r.text())[:500]
    upd={**payload,'ids':[row_id]}
    r=await ctx.request.put('https://api.jbfdigital.com.br/campaigns/Messenger/update-many', headers=h, data=json.dumps(upd), timeout=120000)
    return r.status, (await r.text())[:500]

def load_state():
    if STATE_PATH.exists(): return json.loads(STATE_PATH.read_text(encoding='utf-8'))
    return {'runs':[], 'mixed_2022':{}}

def save_state(state):
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    fd,tmp=tempfile.mkstemp(prefix=STATE_PATH.name+'.', dir=str(STATE_PATH.parent))
    with os.fdopen(fd,'w',encoding='utf-8') as f:
        json.dump(state,f,ensure_ascii=False,indent=2,sort_keys=True); f.write('\n')
    os.replace(tmp, STATE_PATH)

def write_excel(path, rows, summary):
    if not Workbook: return None
    wb=Workbook(); ws=wb.active; ws.title='Paginas'
    headers=['link da pagina','nome da pagina','segurador','bot user','data','codigo dos erros','sb status antes','sb restricted antes','acao','readback ok','observacao']
    ws.append(headers)
    for c in ws[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78'); c.alignment=Alignment(horizontal='center')
    for r in rows:
        ws.append([r.get(h,'') for h in headers])
    for row in range(2,ws.max_row+1):
        cell=ws.cell(row=row,column=1)
        if cell.value and str(cell.value).startswith('http'):
            cell.hyperlink=cell.value; cell.style='Hyperlink'
    for i,w in enumerate([28,24,22,30,18,24,16,18,30,14,35],1): ws.column_dimensions[get_column_letter(i)].width=w
    ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
    sw=wb.create_sheet('Resumo'); sw.append(['Campo','Valor'])
    for k,v in summary.items():
        if isinstance(v,(dict,list)): v=json.dumps(v,ensure_ascii=False)
        sw.append([k,v])
    for c in sw[1]: c.font=Font(bold=True)
    wb.save(path); return str(path)

async def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--apply', action='store_true')
    ap.add_argument('--user', action='append', default=[])
    ap.add_argument('--limit-users', type=int, default=0)
    ap.add_argument('--start-at', default='', help='Resume sorted user list at this bot-user email (inclusive).')
    ap.add_argument('--limit-accounts', type=int, default=0)
    ap.add_argument('--limit-pages', type=int, default=0)
    ap.add_argument('--max-writes', type=int, default=0, help='Canary safety cap for total writes')
    ap.add_argument('--quiet-noop', action='store_true')
    args=ap.parse_args()

    LOG_DIR.mkdir(parents=True, exist_ok=True); REPORT_DIR.mkdir(parents=True, exist_ok=True)
    stamp=datetime.now(NY).strftime('%Y%m%d-%H%M%S'); tday=today()
    run_log=LOG_DIR/f'dtr-sb-page-health-sync-{stamp}.json'
    report_xlsx=REPORT_DIR/f'dtr-sb-page-health-sync-{stamp}.xlsx'
    summary={'ok':True,'mode':'apply' if args.apply else 'dry-run','started_at':now_iso(),'today':tday,'errors':[],'changes':[],'log':str(run_log),'xlsx':str(report_xlsx)}
    p=browser=ctx=None
    try:
        srows=sheet_rows(); active=active_users_from_sheet(srows); summary['sheet_active_users']=len(active)
        matched, missing, op_errors=discover_dtr_items(set(active)); summary['matched_1p_users']=len(matched); summary['missing_1p_users']=missing; summary['op_errors']=op_errors
        users=sorted(matched)
        if args.user:
            req={norm_email(u) for u in args.user}; users=[u for u in users if u in req]; miss=sorted(req-set(users))
            if miss: summary['errors'].append({'missing_requested_users':miss})
        if args.limit_users: users=users[:args.limit_users]
        if args.start_at:
            start=norm_email(args.start_at)
            users=[u for u in users if u >= start]
        p,browser,ctx,h=await get_sb_context(); pubs,sb_rows=await fetch_sb_rows(ctx,h); summary['sb_rows']=len(sb_rows); summary['sb_publishers']=len(pubs)
        indexes=build_sb_indexes(sb_rows)
        sb_restricted_ids={str(r.get('ID')) for r in sb_rows if norm(r.get('STATUS'))=='Broadcast' and active_restricted(r,tday)}
        summary['sb_active_restricted_start']=len(sb_restricted_ids)
        state=load_state(); stats=Counter(); report_rows=[]; backups=[]; writes=0
        for user in users:
            print(f"PROGRESS user_start {user}", flush=True)
            scan=await scan_dtr_user(user, matched[user], args.limit_accounts, args.limit_pages)
            print(f"PROGRESS user_done {user} accounts={len(scan.get('accounts') or [])} reports={len(scan.get('reports') or [])} errors={len(scan.get('errors') or [])}", flush=True)
            stats['users_scanned']+=1; stats['dtr_accounts']+=len(scan.get('accounts') or []); stats['dtr_pages']+=sum(a.get('pages',0) for a in scan.get('accounts') or [])
            if scan.get('errors'): summary['errors'].append({'user':user,'errors':scan['errors']})
            unsafe_context = scan.get('context_signatures_unique',0) < max(1, len(scan.get('accounts') or [])) and len(scan.get('accounts') or [])>1
            if unsafe_context:
                summary['errors'].append({'user':user,'warning':'account_context_signatures_not_unique','unique':scan.get('context_signatures_unique'),'accounts':len(scan.get('accounts') or []),'action':'skipped_automatic_writes'})
                stats['unsafe_context_users'] += 1
            reports = [] if unsafe_context else (scan.get('reports') or [])
            for rep_idx, rep in enumerate(reports, start=1):
                if rep_idx == 1 or rep_idx % 25 == 0 or rep_idx == len(reports):
                    print(f"PROGRESS user_write {user} {rep_idx}/{len(reports)}", flush=True)
                cls=rep['classification']; note=cls.get('note_code') or ''
                status=cls.get('status') or ''; codes=cls.get('codes') or []
                if status=='SENT': stats['sent']+=1
                elif status=='SEM_COMPLETED': stats['sem_completed']+=1
                else: stats['error_pages']+=1
                for c in codes or ([note] if note else []): stats[f'code_{c}']+=1
                sb, merr=match_sb(rep,indexes)
                action=[]; obs=[]; readback_ok=''
                if not sb:
                    stats[merr or 'match_error']+=1; obs.append(merr or 'match_error')
                else:
                    before=public_row(sb); backups.append(before)
                    payload={}
                    # NOTES: every non-Sent result, all statuses.
                    if status!='SENT' and note:
                        new_notes, changed = append_note(sb.get('NOTES'), note)
                        if changed:
                            payload['NOTES']=new_notes; action.append('notes')
                    # Restricted rules.
                    has_2022 = '#2022' in codes
                    is_restricted_start = str(sb.get('ID')) in sb_restricted_ids
                    sb_status = norm(sb.get('STATUS'))
                    if has_2022 and cls.get('restricted_until'):
                        if sb_status == 'On-hold':
                            obs.append('onhold_2022_no_restricted_write')
                            stats['onhold_2022_skipped'] += 1
                        elif sb_status == 'Blocked':
                            fb_status=await fb_page_opens(ctx, norm(sb.get('FB_PAGE_ID')) or norm(rep.get('fb_page_id')))
                            obs.append('fb_'+fb_status); stats[f'blocked_fb_{fb_status}']+=1
                            if fb_status=='available':
                                payload['STATUS']='Broadcast'; payload['RESTRICTED_UNTIL']=cls['restricted_until']; action.append('blocked_to_broadcast'); action.append('restricted_until')
                            else:
                                stats['blocked_2022_restricted_skipped'] += 1
                        else:
                            payload['STATUS']='Broadcast'; payload['RESTRICTED_UNTIL']=cls['restricted_until']; action.append('restricted_until')
                        if len(codes)>1:
                            state.setdefault('mixed_2022',{})[str(sb.get('ID'))]={'last_seen':now_iso(),'codes':codes,'restricted_until':cls['restricted_until'],'sb':before,'dtr':rep,'needs_post_expiry_review':True}
                    elif is_restricted_start and status=='SENT':
                        payload['RESTRICTED_UNTIL']=None; action.append('clear_restricted_sent')
                    elif is_restricted_start and status not in {'SEM_COMPLETED'} and not has_2022:
                        payload['RESTRICTED_UNTIL']=None; action.append('clear_restricted_no2022')
                    # Blocked rule: only Broadcast if FB page opens. If #2022 already handled the blocked check above, do not test twice.
                    if sb_status=='Blocked' and not has_2022:
                        fb_status=await fb_page_opens(ctx, norm(sb.get('FB_PAGE_ID')) or norm(rep.get('fb_page_id')))
                        obs.append('fb_'+fb_status); stats[f'blocked_fb_{fb_status}']+=1
                        if fb_status=='available':
                            payload['STATUS']='Broadcast'; action.append('blocked_to_broadcast')
                    if payload:
                        if args.apply:
                            if args.max_writes and writes>=args.max_writes:
                                obs.append('write_cap_reached')
                            else:
                                st, txt = await sb_update(ctx,h,sb,payload); writes+=1
                                if st not in (200,201):
                                    summary['errors'].append({'update_failed':before,'status':st,'text':txt,'payload':payload}); readback_ok='no'
                                else:
                                    # Fast exact readback; do not refetch all 3,237 rows after every write.
                                    new_sb=await sb_get_row(ctx,h,str(sb.get('ID')))
                                    checks=[]
                                    if new_sb:
                                        if 'NOTES' in payload: checks.append(norm(new_sb.get('NOTES'))==norm(payload['NOTES']))
                                        if 'STATUS' in payload: checks.append(norm(new_sb.get('STATUS'))==norm(payload['STATUS']))
                                        if 'RESTRICTED_UNTIL' in payload: checks.append(date_only(new_sb.get('RESTRICTED_UNTIL'))==date_only(payload['RESTRICTED_UNTIL']))
                                        readback_ok='yes' if all(checks) else 'no'
                                        if readback_ok=='no': summary['errors'].append({'readback_failed':before,'payload':payload,'after':public_row(new_sb)})
                                    else:
                                        readback_ok='no'; summary['errors'].append({'readback_get_failed':before,'payload':payload})
                        else:
                            readback_ok='dry-run'
                        stats['planned_or_done_writes']+=1
                report_rows.append({'link da pagina':('https://facebook.com/'+(norm(rep.get('fb_page_id')) or (norm(sb.get('FB_PAGE_ID')) if sb else ''))) if (norm(rep.get('fb_page_id')) or (norm(sb.get('FB_PAGE_ID')) if sb else '')) else '', 'nome da pagina':rep.get('page_name'), 'segurador':rep.get('account_name'), 'bot user':rep.get('bot_user'), 'data':rep.get('completed_date'), 'codigo dos erros':note or 'Sent', 'sb status antes':norm(sb.get('STATUS')) if sb else '', 'sb restricted antes':date_only(sb.get('RESTRICTED_UNTIL')) if sb else '', 'acao':', '.join(action), 'readback ok':readback_ok, 'observacao':'; '.join(obs)})
                if args.apply and args.max_writes and writes>=args.max_writes:
                    break
            if args.apply and args.max_writes and writes>=args.max_writes: break
        summary['stats']=dict(stats); summary['writes']=writes; summary['backup_rows']=len(backups); summary['finished_at']=now_iso()
        backup_path=REPORT_DIR/f'dtr-sb-page-health-sync-backup-{stamp}.json'
        backup_path.write_text(json.dumps(backups,ensure_ascii=False,indent=2),encoding='utf-8'); summary['backup']=str(backup_path)
        if report_rows: write_excel(report_xlsx, report_rows, summary)
        state.setdefault('runs',[]).append({'ts':summary['started_at'],'mode':summary['mode'],'stats':summary['stats'],'writes':writes,'log':str(run_log),'xlsx':str(report_xlsx)})
        save_state(state)
        if summary['errors']: summary['ok']=False
    except Exception as exc:
        summary['ok']=False; summary['errors'].append({'fatal':f'{type(exc).__name__}: {exc}'})
    finally:
        try:
            if browser: await browser.close()
            if p: await p.stop()
        except Exception: pass
    run_log.write_text(json.dumps(summary,ensure_ascii=False,indent=2,sort_keys=True),encoding='utf-8')
    if args.quiet_noop and not summary.get('writes') and summary.get('ok'):
        return
    print(json.dumps({k:summary.get(k) for k in ['ok','mode','sheet_active_users','matched_1p_users','sb_rows','sb_active_restricted_start','stats','writes','log','xlsx','backup','errors']},ensure_ascii=False,indent=2))
    sys.exit(0 if summary.get('ok') else 2)

if __name__=='__main__': asyncio.run(main())
