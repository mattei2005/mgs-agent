#!/usr/bin/env python3
import json
from datetime import datetime
from pathlib import Path
from collections import Counter

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

base = Path('/root/mgs-agent/reports')
files = {
    'canary_backup': base / 'sb-page-id-fix-backup-20260705-144825.json',
    'canary_result': base / 'sb-page-id-fix-result-canary-20260705-144825.json',
    'apply_backup': base / 'sb-page-id-fix-backup-20260705-144906.json',
    'apply_result': base / 'sb-page-id-fix-result-apply-20260705-144906.json',
    'final_validation': base / 'sb-page-id-fix-plan-20260705-145615.json',
}

def load(p):
    with open(p, encoding='utf-8') as f:
        return json.load(f)

data = {k: load(v) for k, v in files.items()}

def parse_resp(r):
    resp = r.get('response')
    if not resp:
        return {}
    try:
        return json.loads(resp)
    except Exception:
        return {}

backup_apply = {b['target']['sb_id']: b for b in data['apply_backup']['backup']}
backup_canary = {b['target']['sb_id']: b for b in data['canary_backup']['backup']}
validation_by_id = {}
for v in data['canary_result'].get('validations', []) + data['apply_result'].get('validations', []):
    tid = v.get('target', {}).get('sb_id')
    if tid:
        validation_by_id[tid] = v

applied = []
seen = set()
for source, result_key, bmap in [
    ('canário', 'canary_result', backup_canary),
    ('aplicação', 'apply_result', backup_apply),
]:
    for r in data[result_key].get('results', []):
        if not r.get('ok'):
            continue
        sid = r['sb_id']
        if sid in seen:
            continue
        seen.add(sid)
        b = bmap.get(sid) or backup_apply.get(sid) or backup_canary.get(sid)
        if not b:
            continue
        t = b['target']
        before = b.get('live_public', {})
        after = parse_resp(r)
        val = validation_by_id.get(sid, {})
        checks = val.get('checks', {})
        val_after = val.get('after', {})
        def av(k):
            return after.get(k) if after.get(k) is not None else val_after.get(k)
        applied.append({
            'Resultado': 'CORRIGIDO',
            'Execução': source,
            'SB_ID': sid,
            'Usuário bot / SB login': t.get('bot_user') or before.get('USER_LOGIN'),
            'Página DTR': t.get('page_name_dtr'),
            'Página SB antes': before.get('PAGE_NAME') or t.get('page_name_sb'),
            'Segurador DTR': t.get('profile_dtr'),
            'Profile SB antes': before.get('PROFILE_NAME') or t.get('profile_sb'),
            'Diferenças detectadas': ', '.join(t.get('diffs', [])),
            'SB PAGE_ID antes': before.get('PAGE_ID'),
            'DTR PAGE_ID alvo': t.get('target_PAGE_ID'),
            'SB PAGE_ID depois': av('PAGE_ID'),
            'PAGE_ID mudou?': 'SIM' if str(before.get('PAGE_ID', '')) != str(av('PAGE_ID', '')) else 'NÃO',
            'SB FB_PAGE_ID antes': before.get('FB_PAGE_ID'),
            'DTR FB_PAGE_ID alvo': t.get('target_FB_PAGE_ID'),
            'SB FB_PAGE_ID depois': av('FB_PAGE_ID'),
            'FB_PAGE_ID mudou?': 'SIM' if str(before.get('FB_PAGE_ID', '')) != str(av('FB_PAGE_ID', '')) else 'NÃO',
            'SB UTM antes': before.get('UTM_CAMPAIGN'),
            'UTM alvo': t.get('target_UTM_CAMPAIGN'),
            'SB UTM depois': av('UTM_CAMPAIGN'),
            'UTM mudou?': 'SIM' if str(before.get('UTM_CAMPAIGN', '')) != str(av('UTM_CAMPAIGN', '')) else 'NÃO',
            'Status SB': before.get('STATUS'),
            'Company': before.get('COMPANY'),
            'Domain': before.get('DOMAIN'),
            'HTTP status': r.get('status'),
            'Validação PAGE_ID': checks.get('PAGE_ID'),
            'Validação FB_PAGE_ID': checks.get('FB_PAGE_ID'),
            'Validação UTM': checks.get('UTM_CAMPAIGN'),
        })

skipped = []
for s in data['final_validation'].get('skipped_duplicates', []) or data['apply_backup'].get('skipped_duplicates', []):
    rs = s.get('report_sb', {})
    skipped.append({
        'Resultado': 'PULADO - DECISÃO MANUAL',
        'SB_ID': s.get('sb_id'),
        'Usuário DTR': s.get('bot_user'),
        'Usuário SB atual': rs.get('bot_user'),
        'Página DTR': s.get('page_name_dtr'),
        'Página SB': s.get('page_name_sb'),
        'Segurador DTR': s.get('profile_dtr'),
        'Profile SB': s.get('profile_sb'),
        'Diferenças detectadas': ', '.join(s.get('diffs', [])),
        'SB PAGE_ID atual': rs.get('page_id'),
        'DTR PAGE_ID alvo': s.get('target_PAGE_ID'),
        'SB FB_PAGE_ID atual': rs.get('fb_page_id'),
        'DTR FB_PAGE_ID alvo': s.get('target_FB_PAGE_ID'),
        'SB UTM atual': 'pg_' + str(rs.get('page_id')) if rs.get('page_id') else '',
        'UTM alvo': s.get('target_UTM_CAMPAIGN'),
        'Status SB': rs.get('status'),
        'Company': rs.get('company'),
        'Domain': rs.get('domain'),
        'Motivo': s.get('skip_reason'),
    })

already = []
for item in data['final_validation'].get('already_ok', []):
    t = item.get('target', {})
    b = item.get('before', {})
    already.append({
        'SB_ID': t.get('sb_id'),
        'Usuário': t.get('bot_user'),
        'Página': t.get('page_name_dtr'),
        'Profile SB': b.get('PROFILE_NAME'),
        'PAGE_ID validado': b.get('PAGE_ID'),
        'FB_PAGE_ID validado': b.get('FB_PAGE_ID'),
        'UTM validada': b.get('UTM_CAMPAIGN'),
        'Status SB': b.get('STATUS'),
    })

out = base / 'smartbidding-correcao-page-id-utm-20260705.xlsx'
wb = Workbook()
ws = wb.active
ws.title = 'Resumo'
summary = [
    ['Item', 'Valor'],
    ['Arquivo gerado em', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
    ['Correções aplicadas', len(applied)],
    ['Pulados para decisão manual', len(skipped)],
    ['Campos corrigidos/validados', 'PAGE_ID, FB_PAGE_ID, UTM_CAMPAIGN'],
    ['Backup canário', str(files['canary_backup'])],
    ['Backup aplicação', str(files['apply_backup'])],
    ['Resultado canário', str(files['canary_result'])],
    ['Resultado aplicação', str(files['apply_result'])],
    ['Validação final', str(files['final_validation'])],
]
for row in summary:
    ws.append(row)
ws.append([])
ws.append(['Usuário', 'Rows corrigidos'])
for user, n in sorted(Counter(row['Usuário bot / SB login'] for row in applied).items()):
    ws.append([user, n])

def write_sheet(name, rows):
    ws = wb.create_sheet(name)
    if not rows:
        ws.append(['Sem linhas'])
        return ws
    headers = list(rows[0].keys())
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, '') for h in headers])
    fill = PatternFill('solid', fgColor='1F4E78')
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    for i, h in enumerate(headers, 1):
        sample = [len(str(row.get(h, ''))) for row in rows[:300]]
        maxlen = max([len(str(h))] + sample)
        ws.column_dimensions[get_column_letter(i)].width = min(max(maxlen + 2, 10), 45)
    return ws

write_sheet('Correções aplicadas', sorted(applied, key=lambda r: (r['Usuário bot / SB login'], r['Página DTR'] or '')))
write_sheet('Pulados decisao manual', skipped)
write_sheet('Validação final OK', sorted(already, key=lambda r: (r['Usuário'] or '', r['Página'] or '')))

for cell in wb['Resumo'][1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
wb['Resumo'].column_dimensions['A'].width = 35
wb['Resumo'].column_dimensions['B'].width = 110

wb.save(out)
wb2 = openpyxl.load_workbook(out, read_only=True)
print('OK arquivo:', out)
print('sheets:', wb2.sheetnames)
for s in wb2.sheetnames:
    ws = wb2[s]
    print(f'{s}: rows={ws.max_row} cols={ws.max_column}')
print('size_bytes:', out.stat().st_size)
