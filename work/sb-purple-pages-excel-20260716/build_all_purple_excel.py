#!/usr/bin/env python3
import json,re
from collections import Counter
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,Font,PatternFill,Border,Side
from openpyxl.utils import get_column_letter
RUN=Path('/root/mgs-agent/work/sb-purple-pages-excel-20260716');BR=Path('/tmp/sb-ares-purple-recheck.json');PG=RUN/'sb-pages-live.json';OUT=RUN/'paginas-vinculadas-templates-roxos-todos-erros.xlsx'
def msgs(r):
 x=r.get('MESSAGES') or [];return json.loads(x) if isinstance(x,str) else x
def ispurple(m):return int(m.get('INVALID_FORMAT') or 0)>0 or int(m.get('ERROR') or 0)>0
def reasons(ms):
 c=Counter()
 for m in ms:
  rr=m.get('REJECTED_REASON') or {}
  if isinstance(rr,dict):
   for k,v in rr.items():c[str(k)]+=int(v or 0)
 return c
def cat(rs):
 s=' | '.join(rs)
 if 'pages_utility_messaging' in s:return '#200 pages_utility_messaging'
 if 'Application has been deleted' in s:return 'Aplicação deletada'
 if 'Application does not have permission' in s:return 'Aplicação sem permissão'
 return 'Erro genérico'
br=json.loads(BR.read_text())['rows'];target={}
for r in br:
 pm=[m for m in msgs(r) if ispurple(m)]
 if int(r.get('PAGES') or 0)>0 and pm:
  rs=reasons(pm);target[r['ID']]={'template':r['NAME'],'pages':int(r['PAGES']),'purple_messages':len(pm),'category':cat(rs),'reason':' | '.join(rs)}
rows=json.loads(PG.read_text())['rows'];active=[];excluded=[]
for r in rows:
 info=target.get(r.get('BROADCAST_TEMPLATE_ID'))
 if not info:continue
 item={'Segurador':r.get('PROFILE_NAME') or '', 'Página':r.get('PAGE_NAME') or '', 'Nome do template':info['template'], 'Link da página':f"https://facebook.com/{r.get('FB_PAGE_ID')}", 'Usuário do bot':r.get('USER_LOGIN') or r.get('LOGIN') or '', 'Page ID (Facebook)':str(r.get('FB_PAGE_ID') or ''), 'PG / Page ID interno':str(r.get('PAGE_ID') or ''), 'Status SB':r.get('STATUS') or '', 'Categoria do roxo':info['category'], 'Mensagens roxas no template':info['purple_messages'], 'Motivo retornado pela SB':info['reason']}
 (active if item['Status SB'] in {'Broadcast','Campaign'} else excluded).append(item)
active.sort(key=lambda x:(x['Categoria do roxo'],x['Nome do template'],x['Página']));excluded.sort(key=lambda x:(x['Nome do template'],x['Status SB'],x['Página']))
counts=Counter(x['Nome do template'] for x in active)
assert all(counts[x['template']]==x['pages'] for x in target.values()),[(x['template'],x['pages'],counts[x['template']]) for x in target.values() if counts[x['template']]!=x['pages']]
wb=Workbook();ws=wb.active;ws.title='Resumo';dark='1F4E78';purple='7030A0';thin=Side(style='thin',color='D9E1F2')
ws['A1']='Páginas vinculadas a templates com status roxo';ws['A1'].font=Font(size=16,bold=True,color='FFFFFF');ws['A1'].fill=PatternFill('solid',fgColor=dark);ws.merge_cells('A1:D1')
summary=[('Gerado em (ET)',datetime.now(ZoneInfo('America/New_York')).isoformat(timespec='seconds')),('Templates com roxo',len(target)),('Mensagens roxas',sum(x['purple_messages'] for x in target.values())),('Páginas ativas vinculadas',len(active)),('Páginas do subconjunto #200',sum(1 for x in active if x['Categoria do roxo']=='#200 pages_utility_messaging')),('Rows não ativas anexadas',len(excluded))]
for i,(k,v) in enumerate(summary,3):ws.cell(i,1,k).font=Font(bold=True);ws.cell(i,2,v)
ws['A11']='Precisão da atribuição';ws['A11'].font=Font(bold=True,color='FFFFFF');ws['A11'].fill=PatternFill('solid',fgColor=purple);ws.merge_cells('A11:D11')
notes=['O erro roxo é agregado no nível da mensagem/template; a SB não devolve o Page ID causador.','As 162 linhas são todas as páginas Broadcast/Campaign vinculadas aos 10 templates que têm mensagens roxas.','Isso não prova que cada uma das 162 páginas causou o erro; é o universo de páginas potencialmente envolvidas.','O subconjunto #200 pages_utility_messaging contém 17 páginas em seis templates.']
for i,n in enumerate(notes,12):ws.cell(i,1,'• '+n);ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4)
ws['A18']='Categoria';ws['B18']='Templates';ws['C18']='Mensagens roxas';ws['D18']='Páginas ativas'
for c in ws[18]:c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor=dark)
cats=sorted({x['category'] for x in target.values()})
for i,c in enumerate(cats,19):vals=[x for x in target.values() if x['category']==c];ws.cell(i,1,c);ws.cell(i,2,len(vals));ws.cell(i,3,sum(x['purple_messages'] for x in vals));ws.cell(i,4,sum(x['pages'] for x in vals))
for col,w in zip('ABCD',[62,20,22,20]):ws.column_dimensions[col].width=w
def add(name,data,color):
 sh=wb.create_sheet(name);sh.sheet_properties.tabColor=color;headers=list(data[0]) if data else list(active[0]);sh.append(headers)
 for c in sh[1]:c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor=dark);c.alignment=Alignment(horizontal='center',wrap_text=True)
 for r in data:
  sh.append([r[h] for h in headers]);link_col=headers.index('Link da página')+1;c=sh.cell(sh.max_row,link_col);c.hyperlink=c.value;c.style='Hyperlink'
 sh.freeze_panes='A2';sh.auto_filter.ref=f'A1:{get_column_letter(len(headers))}{max(1,sh.max_row)}';widths=[24,28,36,40,22,20,62,14,31,25,58]
 widths=[24,28,62,36,40,22,20,14,31,25,58]
 for i,w in enumerate(widths,1):sh.column_dimensions[get_column_letter(i)].width=w
 for row in sh.iter_rows(min_row=2):
  for c in row:c.alignment=Alignment(vertical='top',wrap_text=True);c.border=Border(bottom=thin)
add('Todas páginas roxas',active,'7030A0');add('Subconjunto #200',[x for x in active if x['Categoria do roxo']=='#200 pages_utility_messaging'],'C000C0');add('Rows não ativas',excluded,'A5A5A5')
wb.save(OUT);rb=load_workbook(OUT);assert rb['Todas páginas roxas'].max_row-1==162;assert rb['Subconjunto #200'].max_row-1==17
print(json.dumps({'output':str(OUT),'templates':len(target),'purple_messages':sum(x['purple_messages'] for x in target.values()),'active_pages':len(active),'subset_200':17,'excluded':len(excluded),'categories':dict(Counter(x['category'] for x in target.values()))},ensure_ascii=False))
