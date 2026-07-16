#!/usr/bin/env python3
import json,re
from collections import Counter
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,Font,PatternFill,Border,Side
from openpyxl.utils import get_column_letter
RUN=Path('/root/mgs-agent/work/sb-purple-pages-excel-20260716')
BROADCAST=Path('/tmp/sb-ares-live-cron-design.json');PAGES=RUN/'sb-pages-live.json';DTR=RUN/'dtr-page-error-scan.json';OUT=RUN/'paginas-roxas-pages-utility-messaging.xlsx'
ERR='(#200) App does not have pages_utility_messaging permission on the Page'
def load_messages(r):
 x=r.get('MESSAGES') or [];return json.loads(x) if isinstance(x,str) else x
def is_target_message(m):
 rr=m.get('REJECTED_REASON') or {};return isinstance(rr,dict) and any('pages_utility_messaging' in str(k) for k in rr)
def dtr_code(text):
 m=re.search(r'error_subcode[^0-9]{0,20}(\d+)',text or '');return m.group(1) if m else ''
br=json.loads(BROADCAST.read_text())['rows'];targets={r['NAME']:{'template_id':r['ID'],'broadcast_pages':int(r.get('PAGES') or 0),'purple_messages':sum(is_target_message(m) for m in load_messages(r))} for r in br if any(is_target_message(m) for m in load_messages(r))}
pages=[r for r in json.loads(PAGES.read_text())['rows'] if r.get('BROADCAST_TEMPLATE_NAME') in targets]
dtr=json.loads(DTR.read_text())['results'];didx={(str(r.get('FB_PAGE_ID')),str(r.get('PAGE_ID'))):r for r in dtr}
active=[];excluded=[]
for r in pages:
 t=targets[r['BROADCAST_TEMPLATE_NAME']];dr=didx.get((str(r.get('FB_PAGE_ID')),str(r.get('PAGE_ID'))),{});sub=dtr_code(dr.get('latest_report_excerpt',''))
 row={'Segurador':r.get('PROFILE_NAME') or '', 'Página':r.get('PAGE_NAME') or '', 'Link da página':f"https://facebook.com/{r.get('FB_PAGE_ID')}", 'Usuário do bot':r.get('USER_LOGIN') or r.get('LOGIN') or '', 'Page ID (Facebook)':str(r.get('FB_PAGE_ID') or ''), 'PG / Page ID interno':str(r.get('PAGE_ID') or ''), 'Template':r.get('BROADCAST_TEMPLATE_NAME') or '', 'Status SB':r.get('STATUS') or '', 'Mensagens roxas com #200':t['purple_messages'], 'Erro SB':ERR, 'Evidência DTR':('OAuthException code 100 / subcode '+sub if sub else dr.get('dtr_result','sem campanha concluída'))}
 (excluded if (r.get('STATUS') or '').casefold()=='on-hold' else active).append(row)
active.sort(key=lambda x:(x['Template'],x['Segurador'],x['Página']));excluded.sort(key=lambda x:x['Página'])
wb=Workbook();ws=wb.active;ws.title='Resumo';dark='1F4E78';blue='D9EAF7';purple='7030A0';green='E2F0D9';gray='E7E6E6';thin=Side(style='thin',color='D9E1F2')
ws['A1']='Páginas sob erro roxo — pages_utility_messaging';ws['A1'].font=Font(size=16,bold=True,color='FFFFFF');ws['A1'].fill=PatternFill('solid',fgColor=dark);ws.merge_cells('A1:D1')
now=datetime.now(ZoneInfo('America/New_York')).isoformat(timespec='seconds');summary=[('Gerado em (ET)',now),('Mensagens roxas com esse motivo',sum(x['purple_messages'] for x in targets.values())),('Templates afetados',len(targets)),('Páginas ativas vinculadas',len(active)),('Rows On-hold fora da contagem do template',len(excluded)),('DTR: páginas ativas com code 100/subcode 1689001',sum('1689001' in x['Evidência DTR'] for x in active))]
for i,(k,v) in enumerate(summary,3):ws.cell(i,1,k).font=Font(bold=True);ws.cell(i,2,v)
ws['A11']='Leitura operacional';ws['A11'].font=Font(bold=True,color='FFFFFF');ws['A11'].fill=PatternFill('solid',fgColor=purple);ws.merge_cells('A11:D11')
notes=['O número 51 representa mensagens roxas, não páginas.','A API de Broadcast agrega o erro por mensagem e não devolve o Page ID responsável.','A aba Páginas afetadas contém as páginas ativas vinculadas aos seis templates com esse erro.','O readback independente no DigitalTRChat encontrou code 100 / subcode 1689001 no último campaign report das 17 páginas ativas.','A aba Fora do escopo contém a row On-hold adicional, não contabilizada pelo PAGES do Broadcast Template.','Use esta lista com o Ciro para validar a permissão pages_utility_messaging/app por página; nenhuma configuração foi alterada nesta auditoria.']
for i,n in enumerate(notes,12):ws.cell(i,1,'• '+n);ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=4);ws.cell(i,1).alignment=Alignment(wrap_text=True)
ws['A20']='Template';ws['B20']='Páginas ativas';ws['C20']='Mensagens roxas #200'
for c in ws[20]:c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor=dark)
ct=Counter(x['Template'] for x in active)
for i,(name,info) in enumerate(sorted(targets.items()),21):ws.cell(i,1,name);ws.cell(i,2,ct[name]);ws.cell(i,3,info['purple_messages'])
ws.column_dimensions['A'].width=62;ws.column_dimensions['B'].width=24;ws.column_dimensions['C'].width=26;ws.column_dimensions['D'].width=15;ws.freeze_panes='A3'
def add_table(name,rows,tabcolor):
 sh=wb.create_sheet(name);sh.sheet_properties.tabColor=tabcolor
 headers=list(rows[0]) if rows else ['Segurador','Página','Link da página','Usuário do bot','Page ID (Facebook)','PG / Page ID interno','Template','Status SB','Mensagens roxas com #200','Erro SB','Evidência DTR']
 sh.append(headers)
 for c in sh[1]:c.font=Font(bold=True,color='FFFFFF');c.fill=PatternFill('solid',fgColor=dark);c.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
 for r in rows:
  sh.append([r[h] for h in headers]);cell=sh.cell(sh.max_row,headers.index('Link da página')+1);cell.hyperlink=cell.value;cell.style='Hyperlink'
 sh.auto_filter.ref=f'A1:{get_column_letter(len(headers))}{max(1,sh.max_row)}';sh.freeze_panes='A2';sh.row_dimensions[1].height=38
 widths={'Segurador':24,'Página':28,'Link da página':36,'Usuário do bot':40,'Page ID (Facebook)':22,'PG / Page ID interno':20,'Template':62,'Status SB':14,'Mensagens roxas com #200':24,'Erro SB':58,'Evidência DTR':42}
 for i,h in enumerate(headers,1):sh.column_dimensions[get_column_letter(i)].width=widths.get(h,20)
 for row in sh.iter_rows(min_row=2):
  for c in row:c.alignment=Alignment(vertical='top',wrap_text=True);c.border=Border(bottom=thin)
 return sh
add_table('Páginas afetadas',active,'C000C0');add_table('Fora do escopo',excluded,'A5A5A5')
OUT.parent.mkdir(parents=True,exist_ok=True);wb.save(OUT)
# Structural readback
rb=load_workbook(OUT,data_only=False);assert rb.sheetnames==['Resumo','Páginas afetadas','Fora do escopo'];assert rb['Páginas afetadas'].max_row-1==17;assert rb['Fora do escopo'].max_row-1==1;assert all(rb['Páginas afetadas'].cell(r,3).hyperlink for r in range(2,rb['Páginas afetadas'].max_row+1))
print(json.dumps({'output':str(OUT),'active_pages':len(active),'excluded_on_hold':len(excluded),'templates':len(targets),'purple_messages':sum(x['purple_messages'] for x in targets.values()),'dtr_1689001':sum('1689001' in x['Evidência DTR'] for x in active),'sheets':rb.sheetnames},ensure_ascii=False))
